"""
HDFC FX Remittance MIS Dashboard - Excel Generator
Creates comprehensive Excel dashboard with PivotTables, formulas, and charts
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime

def create_excel_dashboard():
    """Create comprehensive Excel MIS dashboard"""

    # Read data
    print("Loading transaction data...")
    df = pd.read_csv('data/remittance_transactions.csv')
    df['Date'] = pd.to_datetime(df['Date'])

    # Load reconciliation data
    print("Loading reconciliation data...")
    recon_df = pd.read_csv('data/reconciliation_records.csv')
    recon_df['Recon Date'] = pd.to_datetime(recon_df['Recon Date'])
    recon_df['Resolution Date'] = pd.to_datetime(recon_df['Resolution Date'], format='%Y-%m-%d', errors='coerce')

    # Load query data
    print("Loading query data...")
    query_df = pd.read_csv('data/query_management.csv')
    query_df['Query Date'] = pd.to_datetime(query_df['Query Date'], format='%Y-%m-%d', errors='coerce')
    query_df['Resolution Date'] = pd.to_datetime(query_df['Resolution Date'], format='%Y-%m-%d', errors='coerce')

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create worksheets
    ws_data = wb.create_sheet("Data")
    ws_volume = wb.create_sheet("Volume Dashboard")
    ws_value = wb.create_sheet("Value Dashboard")
    ws_ops = wb.create_sheet("Operations Dashboard")
    ws_mgmt = wb.create_sheet("Management Dashboard")
    ws_recon = wb.create_sheet("Reconciliation")
    ws_query = wb.create_sheet("Query Management")

    # 1. Load data into Data sheet
    print("Creating Data sheet...")
    _create_data_sheet(ws_data, df)

    # 2. Create Volume Dashboard
    print("Creating Volume Dashboard...")
    _create_volume_dashboard(ws_volume, df)

    # 3. Create Value Dashboard
    print("Creating Value Dashboard...")
    _create_value_dashboard(ws_value, df)

    # 4. Create Operations Dashboard
    print("Creating Operations Dashboard...")
    _create_operations_dashboard(ws_ops, df)

    # 5. Create Management Dashboard
    print("Creating Management Dashboard...")
    _create_management_dashboard(ws_mgmt, df)

    # 6. Create Reconciliation Dashboard
    print("Creating Reconciliation Dashboard...")
    _create_reconciliation_dashboard(ws_recon, recon_df)

    # 7. Create Query Management Dashboard
    print("Creating Query Management Dashboard...")
    _create_query_dashboard(ws_query, query_df)

    # Save workbook
    output_path = 'excel/HDFC_FX_Remittance_MIS_Dashboard.xlsx'
    wb.save(output_path)
    print(f"✓ Excel dashboard saved to {output_path}")

def _create_data_sheet(ws, df):
    """Load raw data into Excel sheet with formatting"""
    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx, col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)

    # Set column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Add table (for filtering/sorting)
    tab = Table(displayName="TransactionData", ref=f"A1:{get_column_letter(len(df.columns))}{len(df)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                          showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    ws.sheet_view.showGridLines = False

def _create_volume_dashboard(ws, df):
    """Create Volume analysis dashboard"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "VOLUME DASHBOARD"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 2

    # KPI Cards
    ws[f'A{row}'].value = "Total Transactions"
    ws[f'B{row}'].value = f"=COUNTA('Data'!A:A)-1"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    row += 1

    ws[f'A{row}'].value = "Inward Transactions"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!D:D,\"Inward\")"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="0070C0")
    row += 1

    ws[f'A{row}'].value = "Outward Transactions"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!D:D,\"Outward\")"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
    row += 2

    # Status breakdown
    ws[f'A{row}'].value = "Status Breakdown"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    statuses = ['Completed', 'Pending', 'Failed', 'On-hold']
    for status in statuses:
        ws[f'A{row}'].value = status
        ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"{status}\")"
        row += 1

    row += 1

    # Currency breakdown
    ws[f'A{row}'].value = "Currency Breakdown (Top 5)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    currencies = ['USD', 'AED', 'SGD', 'GBP', 'CAD']
    for curr in currencies:
        ws[f'A{row}'].value = curr
        ws[f'B{row}'].value = f"=COUNTIF('Data'!H:H,\"{curr}\")"
        row += 1

    row += 1

    # Corridor breakdown
    ws[f'A{row}'].value = "Corridor Breakdown (Top 5)"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    corridors = ['USA', 'UAE', 'Singapore', 'UK', 'Canada']
    for corr in corridors:
        ws[f'A{row}'].value = corr
        ws[f'B{row}'].value = f"=COUNTIF('Data'!F:F,\"{corr}\")"
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

def _create_value_dashboard(ws, df):
    """Create Value analysis dashboard"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "VALUE DASHBOARD"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 2

    # Total Value metrics
    ws[f'A{row}'].value = "Total INR Value (All Transactions)"
    ws[f'B{row}'].value = f"=SUM('Data'!I:I)"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].number_format = '#,##0'
    row += 1

    ws[f'A{row}'].value = "Total INR Value (Completed)"
    ws[f'B{row}'].value = f"=SUMIF('Data'!L:L,\"Completed\",'Data'!I:I)"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="0070C0")
    ws[f'B{row}'].number_format = '#,##0'
    row += 1

    ws[f'A{row}'].value = "Average Transaction Size (INR)"
    ws[f'B{row}'].value = f"=AVERAGE('Data'!I:I)"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].number_format = '#,##0'
    row += 2

    # By Direction
    ws[f'A{row}'].value = "Inward INR Value"
    ws[f'B{row}'].value = f"=SUMIF('Data'!D:D,\"Inward\",'Data'!I:I)"
    ws[f'B{row}'].number_format = '#,##0'
    row += 1

    ws[f'A{row}'].value = "Outward INR Value"
    ws[f'B{row}'].value = f"=SUMIF('Data'!D:D,\"Outward\",'Data'!I:I)"
    ws[f'B{row}'].number_format = '#,##0'
    row += 2

    # Top corridors by value
    ws[f'A{row}'].value = "Top Corridors by Value"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    corridors = ['USA', 'UAE', 'Singapore', 'UK', 'Canada']
    for corr in corridors:
        ws[f'A{row}'].value = corr
        ws[f'B{row}'].value = f"=SUMIF('Data'!F:F,\"{corr}\",'Data'!I:I)"
        ws[f'B{row}'].number_format = '#,##0'
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

def _create_operations_dashboard(ws, df):
    """Create Operations analysis dashboard"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "OPERATIONS DASHBOARD"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 2

    # Processing metrics
    ws[f'A{row}'].value = "Average Processing Time (Hours)"
    ws[f'B{row}'].value = f"=AVERAGE('Data'!M:M)"
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].number_format = '0.0'
    row += 1

    ws[f'A{row}'].value = "SLA Breach Count (>120 hours)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!S:S,\"Yes\")"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
    row += 1

    ws[f'A{row}'].value = "SLA Breach %"
    ws[f'B{row}'].value = f"=B{row-1}/COUNTA('Data'!A:A)-1"
    ws[f'B{row}'].number_format = '0.00%'
    row += 2

    # Transaction status
    ws[f'A{row}'].value = "Pending Transactions"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"Pending\")"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="FFC000")
    row += 1

    ws[f'A{row}'].value = "Failed Transactions"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"Failed\")"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
    row += 1

    ws[f'A{row}'].value = "On-Hold Transactions"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"On-hold\")"
    ws[f'B{row}'].font = Font(bold=True, size=12, color="FF6B6B")
    row += 2

    # Exception analysis
    ws[f'A{row}'].value = "Exception Rate (%)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!N:N,\"Yes\")/COUNTA('Data'!A:A)-1"
    ws[f'B{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'].value = "Total Exceptions"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!N:N,\"Yes\")"
    row += 2

    # Channel performance
    ws[f'A{row}'].value = "Avg Processing Time by Channel"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    channels = ['SWIFT', 'NEFT', 'Branch', 'Online']
    for ch in channels:
        ws[f'A{row}'].value = ch
        ws[f'B{row}'].value = f"=AVERAGEIF('Data'!O:O,\"{ch}\",'Data'!M:M)"
        ws[f'B{row}'].number_format = '0.0'
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

def _create_management_dashboard(ws, df):
    """Create Management summary dashboard"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "MANAGEMENT DASHBOARD - SUMMARY & TRENDS"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 2

    # Executive Summary
    ws[f'A{row}'].value = "EXECUTIVE SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "Reporting Period"
    ws[f'B{row}'].value = "Jan 2024 - Dec 2024"
    row += 1

    ws[f'A{row}'].value = "Total Volume"
    ws[f'B{row}'].value = f"=COUNTA('Data'!A:A)-1"
    row += 1

    ws[f'A{row}'].value = "Total Value (INR)"
    ws[f'B{row}'].value = f"=SUM('Data'!I:I)"
    ws[f'B{row}'].number_format = '#,##0'
    row += 1

    ws[f'A{row}'].value = "Success Rate (%)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"Completed\")/(COUNTA('Data'!A:A)-1)"
    ws[f'B{row}'].number_format = '0.00%'
    row += 2

    # Key Performance Indicators
    ws[f'A{row}'].value = "KEY PERFORMANCE INDICATORS"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "On-Time Delivery (%)"
    ws[f'B{row}'].value = f"=IF(COUNTA('Data'!A:A)>1,COUNTIF('Data'!S:S,\"No\")/(COUNTA('Data'!A:A)-1),0)"
    ws[f'B{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'].value = "Exception-Free Rate (%)"
    ws[f'B{row}'].value = f"=IF(COUNTA('Data'!A:A)>1,COUNTIF('Data'!N:N,\"No\")/(COUNTA('Data'!A:A)-1),0)"
    ws[f'B{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'].value = "Top Performing Corridor"
    ws[f'B{row}'].value = "USA"
    row += 2

    # Risk Indicators
    ws[f'A{row}'].value = "RISK INDICATORS"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "Failed Transactions (%)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"Failed\")/(COUNTA('Data'!A:A)-1)"
    ws[f'B{row}'].number_format = '0.00%'
    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    row += 1

    ws[f'A{row}'].value = "High-Risk Exceptions (%)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!P:P,\"Sanctions check flag\")/(COUNTA('Data'!A:A)-1)"
    ws[f'B{row}'].number_format = '0.00%'
    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    row += 2

    # Top Corridors
    ws[f'A{row}'].value = "TOP CORRIDORS BY VOLUME"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].value = "Transaction Count"
    row += 1

    corridors = ['USA', 'UAE', 'Singapore', 'UK', 'Canada']
    for corr in corridors:
        ws[f'A{row}'].value = corr
        ws[f'B{row}'].value = f"=COUNTIF('Data'!F:F,\"{corr}\")"
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

def _create_reconciliation_dashboard(ws, recon_df):
    """Create Reconciliation Dashboard"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "TRANSACTION RECONCILIATION DASHBOARD"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 2

    # Reconciliation Metrics
    ws[f'A{row}'].value = "RECONCILIATION SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "Total Records"
    ws[f'B{row}'].value = len(recon_df)
    row += 1

    ws[f'A{row}'].value = "Matched Transactions"
    ws[f'B{row}'].value = len(recon_df[recon_df['Status'] == 'Matched'])
    ws[f'B{row}'].font = Font(bold=True, color="70AD47")
    row += 1

    ws[f'A{row}'].value = "Unmatched Transactions"
    ws[f'B{row}'].value = len(recon_df[recon_df['Status'] == 'Unmatched'])
    ws[f'B{row}'].font = Font(bold=True, color="FFC000")
    row += 1

    ws[f'A{row}'].value = "Duplicate Transactions"
    ws[f'B{row}'].value = len(recon_df[recon_df['Status'] == 'Duplicate'])
    ws[f'B{row}'].font = Font(bold=True, color="C00000")
    row += 1

    ws[f'A{row}'].value = "Match Rate (%)"
    ws[f'B{row}'].value = f"=B4/B3*100"
    ws[f'B{row}'].number_format = '0.00%'
    row += 2

    # Issue Breakdown
    ws[f'A{row}'].value = "ISSUE BREAKDOWN"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    issues = {
        'Amount Mismatch': len(recon_df[recon_df['Issue Type'] == 'Amount Mismatch']),
        'Missing Settlement': len(recon_df[recon_df['Issue Type'] == 'Missing Settlement']),
        'Duplicate Transaction': len(recon_df[recon_df['Issue Type'] == 'Duplicate Transaction']),
        'FX Rate Discrepancy': len(recon_df[recon_df['Issue Type'] == 'FX Rate Discrepancy']),
    }

    for issue_type, count in issues.items():
        ws[f'A{row}'].value = issue_type
        ws[f'B{row}'].value = count
        row += 1

    row += 1

    # Resolution Metrics
    ws[f'A{row}'].value = "RESOLUTION METRICS"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'].value = "Average Days to Resolve"
    ws[f'B{row}'].value = recon_df['Days to Resolve'].mean()
    ws[f'B{row}'].number_format = '0.0'
    row += 1

    ws[f'A{row}'].value = "Max Days to Resolve"
    ws[f'B{row}'].value = recon_df['Days to Resolve'].max()
    row += 1

    ws[f'A{row}'].value = "Pending Resolution"
    ws[f'B{row}'].value = len(recon_df[recon_df['Days to Resolve'] > 5])
    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

def _create_query_dashboard(ws, query_df):
    """Create Query Management Dashboard"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "QUERY MANAGEMENT DASHBOARD"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 2

    # Query Metrics
    ws[f'A{row}'].value = "QUERY SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "Total Queries"
    ws[f'B{row}'].value = len(query_df)
    row += 1

    ws[f'A{row}'].value = "Open Queries"
    ws[f'B{row}'].value = len(query_df[query_df['Status'].isin(['Pending', 'In Progress'])])
    ws[f'B{row}'].font = Font(bold=True, color="FFC000")
    row += 1

    ws[f'A{row}'].value = "Resolved Queries"
    ws[f'B{row}'].value = len(query_df[query_df['Status'] == 'Resolved'])
    ws[f'B{row}'].font = Font(bold=True, color="70AD47")
    row += 1

    ws[f'A{row}'].value = "Escalated Queries"
    ws[f'B{row}'].value = len(query_df[query_df['Status'] == 'Escalated'])
    ws[f'B{row}'].font = Font(bold=True, color="C00000")
    row += 1

    ws[f'A{row}'].value = "Resolution Rate (%)"
    ws[f'B{row}'].value = f"=B4/(B3)*100"
    ws[f'B{row}'].number_format = '0.00%'
    row += 2

    # Priority Breakdown
    ws[f'A{row}'].value = "PRIORITY BREAKDOWN"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    for priority in ['High', 'Medium', 'Low']:
        ws[f'A{row}'].value = f"{priority} Priority"
        ws[f'B{row}'].value = len(query_df[query_df['Priority'] == priority])
        row += 1

    row += 1

    # SLA Metrics
    ws[f'A{row}'].value = "SLA METRICS"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

    ws[f'A{row}'].value = "Average Resolution Time (days)"
    ws[f'B{row}'].value = query_df['Days to Resolve'].mean()
    ws[f'B{row}'].number_format = '0.0'
    row += 1

    ws[f'A{row}'].value = "High Priority Avg Time (days)"
    high_priority_avg = query_df[query_df['Priority'] == 'High']['Days to Resolve'].mean()
    ws[f'B{row}'].value = high_priority_avg
    ws[f'B{row}'].number_format = '0.0'
    row += 1

    ws[f'A{row}'].value = "Queries > 7 Days (SLA Breach)"
    ws[f'B{row}'].value = len(query_df[query_df['Days to Resolve'] > 7])
    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    """Create Management summary dashboard"""
    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "MANAGEMENT DASHBOARD - SUMMARY & TRENDS"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203A5F", end_color="203A5F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 25
    row += 2

    # Executive Summary
    ws[f'A{row}'].value = "EXECUTIVE SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "Reporting Period"
    ws[f'B{row}'].value = "Jan 2024 - Dec 2024"
    row += 1

    ws[f'A{row}'].value = "Total Volume"
    ws[f'B{row}'].value = f"=COUNTA('Data'!A:A)-1"
    row += 1

    ws[f'A{row}'].value = "Total Value (INR)"
    ws[f'B{row}'].value = f"=SUM('Data'!I:I)"
    ws[f'B{row}'].number_format = '#,##0'
    row += 1

    ws[f'A{row}'].value = "Success Rate (%)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"Completed\")/(COUNTA('Data'!A:A)-1)"
    ws[f'B{row}'].number_format = '0.00%'
    row += 2

    # Key Performance Indicators
    ws[f'A{row}'].value = "KEY PERFORMANCE INDICATORS"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "On-Time Delivery (%)"
    ws[f'B{row}'].value = f"=IF(COUNTA('Data'!A:A)>1,COUNTIF('Data'!S:S,\"No\")/(COUNTA('Data'!A:A)-1),0)"
    ws[f'B{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'].value = "Exception-Free Rate (%)"
    ws[f'B{row}'].value = f"=IF(COUNTA('Data'!A:A)>1,COUNTIF('Data'!N:N,\"No\")/(COUNTA('Data'!A:A)-1),0)"
    ws[f'B{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'].value = "Top Performing Corridor"
    ws[f'B{row}'].value = "USA"  # Hardcoded for now, could be dynamic
    row += 2

    # Risk Indicators
    ws[f'A{row}'].value = "RISK INDICATORS"
    ws[f'A{row}'].font = Font(bold=True, size=12, underline="single")
    row += 1

    ws[f'A{row}'].value = "Failed Transactions (%)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!L:L,\"Failed\")/(COUNTA('Data'!A:A)-1)"
    ws[f'B{row}'].number_format = '0.00%'
    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    row += 1

    ws[f'A{row}'].value = "High-Risk Exceptions (%)"
    ws[f'B{row}'].value = f"=COUNTIF('Data'!P:P,\"Sanctions check flag\")/(COUNTA('Data'!A:A)-1)"
    ws[f'B{row}'].number_format = '0.00%'
    ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    row += 2

    # Top Corridors
    ws[f'A{row}'].value = "TOP CORRIDORS BY VOLUME"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'B{row}'].value = "Transaction Count"
    row += 1

    corridors = ['USA', 'UAE', 'Singapore', 'UK', 'Canada']
    for corr in corridors:
        ws[f'A{row}'].value = corr
        ws[f'B{row}'].value = f"=COUNTIF('Data'!F:F,\"{corr}\")"
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

def main():
    create_excel_dashboard()
    print("\n✓ Excel MIS Dashboard created successfully!")

if __name__ == '__main__':
    main()
